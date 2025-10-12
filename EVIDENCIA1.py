
from datetime import datetime, timedelta
salas = {}
id_sala = 1000

clientes = {}
id_cliente = 0

reservas = {}
folio_contador = 0


while True:
  print('\n************ MENÚ ************')
  print('1. REGISTRAR RESERVACION DE UNA SALA')
  print('2. EDITAR EL NOMBRE DE UN EVENTO DE UNA RESERVACION YA HECHA')
  print('3. CONSULTAR RESERVACIONES EXISTENTES')
  print('4. REGISTRAR UN CLIENTE')
  print('5. REGISTRAR UNA SALA')
  print('6. SALIR') 

  opcion = input('Ingrese una opción: ')


  if opcion not in '123456':
     print('\nError. Respuesta no valida')
     continue
  

  if opcion == '1':
    if clientes:
      while True:

        clientes_ordenados = sorted(
        clientes.items(),
        key=lambda item: item[1][1].lower()
        )

        print("\n********** Clientes Registrados **********")
        print(f'{'ID':<5}{'Apellido':<20}{'Nombre':<20}')
        for id, datos in clientes_ordenados:
            nombre, apellido = datos
            print(f'{id:<5}{apellido:<20}{nombre:<20}')


        id_cliente_buscar = input('\nDigite su ID ("X" para salir): ').strip().upper()
        if id_cliente_buscar ==  'X':
          print('Saliendo...')
          break
        if not id_cliente_buscar.isdigit():
          print('Error. Id invalido')
          continue
        else:
          id_cliente_buscar = int(id_cliente_buscar)
          if id_cliente_buscar not in clientes.keys():
            print('Error. ID no registrado')
            continue
          else:
            print('ID encontrado')


        hoy = datetime.now().date()  
        fecha_minima = hoy + timedelta(days=2)
        while True:
          fecha_usuario_str=input("\nINGRESE LA FECHA DE RESERVA (DD-MM-AAAA): ").strip()
          try:
              fecha_usuario = datetime.strptime(fecha_usuario_str, "%d-%m-%Y").date()

              if fecha_usuario >= fecha_minima:
                print("La fecha es valida para reserva")
                break
              else:
                print(f'Error. La fecha debe de ser por lo menos 2 dias posteriores a hoy {hoy}')
                continue
          except ValueError:
            print("Formato de fecha invalido")
            continue
        

        turnos = {'M':'Matutino', 'V': 'Vespertino', 'N': 'Nocturno'}
        while True:
          print('TURNOS: "Matutino"(M) | "Vespertino"(V) | "Nocturno"(N)')
          seleccionar_turno = input("Ingrese turno el turno que desea (M/V/N): ").upper().strip()
          if seleccionar_turno not in turnos.keys():
            print('Error. Seleccione un turno valido')
            continue
          else:
            turno_sala = turnos[seleccionar_turno]
            print('Turno registrado')
            break
        
        print("\n********** SALAS DISPONIBLES **********")
        print(f'{'ID':<5}{'Nombre':<15}{'Cupo':<15}')
        salas_disponibles = []
        for id_sala_item, (nombre_sala, cupo_sala) in salas.items():
          ocupada = False
          for folio, datos_reserva in reservas.items():
            if (datos_reserva['sala_id'] == id_sala_item and
                datos_reserva['fecha'] == fecha_usuario and
                datos_reserva['turno'] == turno_sala):
              ocupada = True
              break
          if not ocupada:
            print(f"{id_sala_item:<5}{nombre_sala:<15}{cupo_sala:<15}")
            salas_disponibles.append(id_sala_item)

        if not salas_disponibles:
          print("No hay salas disponibles en esa fecha y turno.")
          break

        while True:
          try:
            sala_seleccionada = int(input("\nIngrese el ID de la sala: "))

            if sala_seleccionada not in salas_disponibles:
              print("La sala seleccionada no está disponible.")
              continue
            else:
              break
          except ValueError:
            print("Error. Debe ingresar un número de sala válido.")
            continue

          

        while True:
          nombre_reserva = input('\nIngrese el nombre de la reservación de la sala: ').strip()
          if not nombre_reserva:
            print('Error. El nombre no puede estar vacío')
            continue
          else:
            break

        
        folio_contador += 1
        folio = f"F{folio_contador}"

        reservas[folio] = {
          'cliente_id': id_cliente_buscar,
          'sala_id': sala_seleccionada,
          'fecha': fecha_usuario,
          'turno': turno_sala,
          'evento': nombre_reserva
        }

        print(f"\nReservación registrada con éxito. Folio: {folio}")
        print(f"Cliente: {clientes[id_cliente_buscar][1]} {clientes[id_cliente_buscar][0]}")
        print(f"Sala: {salas[sala_seleccionada][0]} (Cupo: {salas[sala_seleccionada][1]})")
        print(f"Fecha: {fecha_usuario} | Turno: {turno_sala} | Evento: {nombre_reserva}")
        break
   

    else:
      print('Error. Aún no hay clientes registrados')
      continue


  if opcion == '2':
    
    if reservas:

    
      try:
        fecha_inicio_str = input("\nIngrese fecha de inicio (DD-MM-AAAA) ('X' para salir): ").strip().upper()
        if fecha_inicio_str == 'X':
          print('Saliendo...')
          continue
        fecha_fin_str = input("Ingrese fecha final (DD-MM-AAAA): ").strip()
        
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%d-%m-%Y").date()
        fecha_fin = datetime.strptime(fecha_fin_str, "%d-%m-%Y").date()

        if fecha_fin < fecha_inicio:
          print("Error. La fecha final no puede ser anterior a la fecha de inicio.")
          continue
      except ValueError:
        print("Formato de fecha inválido")
        continue

    
      reservas_en_rango = {}
      for folio, datos in reservas.items():
        if fecha_inicio <= datos['fecha'] <= fecha_fin:
          reservas_en_rango[folio] = datos

      if not reservas_en_rango:
        print("No hay reservaciones en ese rango de fechas.")
        continue

    
      print(f'\n{'FOLIO'}\t{'FECHA'}\t\t{'SALA'}\t\t{'CLIENTE'}\t\t{'EVENTO'}\t\t{'TURNO'}')
      for folio, datos in reservas_en_rango.items():
        cliente_id = datos['cliente_id']
        sala_id = datos['sala_id']
        nombre_cliente = f"{clientes[cliente_id][1]} {clientes[cliente_id][0]}"
        nombre_sala = salas[sala_id][0]
        print(f'{folio}\t{datos['fecha']}\t{nombre_sala}\t\t{nombre_cliente}\t{datos['evento']}\t\t{datos['turno']}')

      
      folio_sel = input("\nIngrese el folio de la reserva que desea modificar: ").strip().upper()
      if folio_sel not in reservas_en_rango.keys():
        print("Error. Folio no válido o fuera del rango.")
        continue

      
      nuevo_evento = input("Ingrese el nuevo nombre del evento: ").strip()
      if not nuevo_evento:
        print("Error. El nombre del evento no puede estar vacío.")
        continue

      reservas[folio_sel]['evento'] = nuevo_evento
      print(f"Reserva {folio_sel} actualizada con éxito. Nuevo evento: {nuevo_evento}")

    else:
      print('No hay reservaciones registradas')
      continue


  if opcion == '3':
    if not reservas:
        print("No hay reservaciones registradas.")
        continue

    
    fechas_con_reservas = sorted({datos['fecha'] for datos in reservas.values()})

    
    print("\nFechas con reservaciones registradas:")
    for fecha in fechas_con_reservas:
        print(f"- {fecha.strftime('%d-%m-%Y')}")

    
    fecha_consulta_str = input("\nIngrese la fecha a consultar (DD-MM-AAAA) ('X' para salir): ").strip()
    
    if fecha_consulta_str.upper() == 'X':
        print('Saliendo...')
        continue

    try:
        fecha_consulta = datetime.strptime(fecha_consulta_str, "%d-%m-%Y").date()
    except ValueError:
        print("Error. Formato de fecha inválido")
        continue

    reservas_para_fecha = {
        folio: datos for folio, datos in reservas.items()
        if datos['fecha'] == fecha_consulta
    }

    if not reservas_para_fecha:
        print(f"No hay reservaciones hechas para la fecha {fecha_consulta}.")
        continue

    print("\n**************************")
    print(f"*  REPORTE DE RESERVAS PARA EL DÍA {fecha_consulta.strftime('%d-%m-%Y')}  *")
    print("**************************")
    print(f"{'FOLIO':<8}{'SALA':<15}{'CLIENTE':<25}{'EVENTO':<25}{'TURNO':<10}")
    for folio, datos in reservas_para_fecha.items():
        cliente_id = datos['cliente_id']
        sala_id = datos['sala_id']
        nombre_cliente = f"{clientes[cliente_id][1]} {clientes[cliente_id][0]}"
        nombre_sala = salas[sala_id][0]
        print(f"{folio:<8}{nombre_sala:<15}{nombre_cliente:<25}{datos['evento']:<25}{datos['turno']:<10}")
    print("********* FIN DEL REPORTE ************")
    
    opcion_exportar = input("\n¿Desea exportar este reporte? (S/N): ").strip().upper()
    if opcion_exportar == 'S':
      print("\nSeleccione formato de exportación:")
      print("1. CSV")
      print("2. JSON")
      print("3. Excel (.xlsx)")
      formato = input("Ingrese una opción (1-3): ").strip()
      
      reporte_export = []
      for folio, datos in reservas_para_fecha.items():
        cliente_id = datos['cliente_id']
        sala_id = datos['sala_id']
        nombre_cliente = f"{clientes[cliente_id][1]} {clientes[cliente_id][0]}"
        nombre_sala = salas[sala_id][0]
        reporte_export.append({
            'Folio': folio,
            'Sala': nombre_sala,
            'Cliente': nombre_cliente,
            'Evento': datos['evento'],
            'Turno': datos['turno']
        })
        
        if formato == '1':
          import csv
          with open("reporte_reservas.csv", mode="w", newline="", encoding="latin1") as archivo:
            writer = csv.DictWriter(archivo, fieldnames=reporte_export[0].keys())
            writer.writeheader()
            writer.writerows(reporte_export)
          print("Reporte exportado a reporte_reservas.csv")
          
        elif formato == '2':
          import json
          with open("reporte_reservas.json", mode="w", encoding="latin1") as archivo:
           json.dump(reporte_export, archivo, indent=4, ensure_ascii=False)
          print("Reporte exportado a reporte_reservas.json")
        
        elif formato == '3':
            from openpyxl import Workbook
            from openpyxl.styles import Font, Border, Side, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Reservas"

    # Título
            ws.merge_cells('A1:E1')
            ws['A1'] = f"REPORTE DE RESERVAS - {fecha_consulta.strftime('%d-%m-%Y')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

    # Encabezados
            columnas = list(reporte_export[0].keys())
            for col, nombre_col in enumerate(columnas, start=1):
                celda = ws.cell(row=2, column=col, value=nombre_col)
                celda.font = Font(bold=True)
                celda.border = Border(bottom=Side(style='thick'))
                celda.alignment = Alignment(horizontal='center')

    # Datos
            for fila, item in enumerate(reporte_export, start=3):
                for col, valor in enumerate(item.values(), start=1):
                    celda = ws.cell(row=fila, column=col, value=valor)
                    celda.alignment = Alignment(horizontal='center')

    # Ajustar ancho de columnas automáticamente
            for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = max_length + 2

    # Guardar archivo 
            nombre_archivo = "reporte_reservas.xlsx"
            wb.save(nombre_archivo)
            print(f"Reporte exportado como '{nombre_archivo}'")
        else:
            print("Opcion no valida. No se exporto el reporte")

  if opcion == '4':
    nombre_cliente = input('\nIngrese su nombre ("X" para salir): ').strip().title()
    if nombre_cliente == 'X':
      print('Saliendo...')
      continue
    apellido_cliente = input('Ingrese su apellido: ').strip().title()
    
    if nombre_cliente and apellido_cliente:
      id_cliente+=1
      clientes[id_cliente] = nombre_cliente, apellido_cliente
      print('Cliente registrado')
    else:
      print('Error. El nombre y apellido no pueden estar vacios') 


  if opcion == '5':
    nombre_sala = input('\nIngrese el nombre de su sala ("X" para salir): ').strip().title()
    if nombre_sala == 'X':
      print('Saliendo...')
      continue
    cupo_sala_str = input('Ingrese el cupo de su sala: ').strip() 
    if nombre_sala and cupo_sala_str:
      try:
        cupo_sala  = int(cupo_sala_str)
        if cupo_sala <= 0:
          raise ValueError
      except ValueError:
        print('Error. Cupo de sala invalido')
      else:   
        id_sala+=1  
        salas[id_sala] = nombre_sala, cupo_sala
        print('Sala creada')
    else:
      print('Error. El nombre y cupo no pueden estar vacios')


  if opcion == '6':
      confirmar = input("¿Está seguro que desea salir? (S/N): ").strip().upper()
      if confirmar == 'S':
          import json
        
        # Serializar clientes
          with open("clientes.json", "w", encoding="utf-8") as archivo:
              json.dump(clientes, archivo, indent=4, ensure_ascii=False)
        
        # Serializar salas
          with open("salas.json", "w", encoding="utf-8") as archivo:
              json.dump(salas, archivo, indent=4, ensure_ascii=False)
        
        # Serializar reservas (fechas como string para JSON)
          reservas_serializables = {
              folio: {
                  **datos,
                  'fecha': datos['fecha'].strftime("%Y-%m-%d")
              }
              for folio, datos in reservas.items()
          }
          with open("reservas.json", "w", encoding="utf-8") as archivo:
              json.dump(reservas_serializables, archivo, indent=4, ensure_ascii=False)
        
          print("Datos guardados correctamente. Gracias por su visita.")
          break
      elif confirmar == 'N':
          print("Salida cancelada. Regresando al menú principal...")
          continue
      else:
          print("respuesta invalida. Por favor ingrese 'S' para salir o 'N' para cancelar.")
          continue
